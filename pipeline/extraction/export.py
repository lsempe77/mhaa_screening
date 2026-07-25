"""
export.py — Turn reconciled extractions into review artifacts.

Reads the extraction JSONL produced by extract.py and writes:

  1. A long-form JSONL  (one row per study x field) — the machine-readable
     master, with full provenance (value, quotes, location, agreement, flags).
  2. An Excel workbook for reviewers:
       - Index          : one row per study (coverage + #cells needing a human)
       - <group> sheets : one row per study, columns per field (value/quotes/
                          location/agreement), conflict cells highlighted
       - Review queue   : only the cells a human must check (conflicts + quote
                          failures) — the reviewer's worklist
       - Eval crosswalk : evaluation section -> contributing fields -> coverage
       - Codebook       : the framework field registry (self-documenting)

Usage:
    python pipeline/extraction/export.py \
        --project girl_effect \
        --extraction projects/girl_effect/full_text/extraction/output/extraction_v1.jsonl \
        --records projects/girl_effect/full_text/data/records_388.jsonl
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

_HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(_HERE.parent))

import config
from framework import Field, Framework, load_framework
# Reuse the screening engine's quote machinery, but drive it with a per-study
# pre-normalised haystack (verify_quote normalises the full 400k-char source on
# every call — far too slow across ~20k consensus quotes).
from k5_runner import norm, parse_quote_fragments, _fuzzy_in_haystack  # type: ignore

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Highlights for cells that need human attention.
FILL_CONFLICT = PatternFill("solid", fgColor="FFF2CC")   # amber — conflict/needs_human
FILL_QUOTEFAIL = PatternFill("solid", fgColor="F8CBAD")  # salmon — unverified quote
FILL_HEADER = PatternFill("solid", fgColor="1F4E78")     # dark blue header
FONT_HEADER = Font(bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")


# --------------------------- value rendering ---------------------------

def render_value(field: Field, value) -> str:
    if value is None or value == "":
        return ""
    if field.type == "boolean":
        return "TRUE" if value else "FALSE"
    if field.type == "categorical_multi":
        return ", ".join(str(v) for v in value) if isinstance(value, list) else str(value)
    if field.type == "composite":
        if isinstance(value, dict):
            return " | ".join(f"{k}: {v}" for k, v in value.items()
                              if v not in (None, "", "null"))
        return str(value)
    if field.type == "composite_multi":
        if isinstance(value, list):
            rows = []
            for item in value:
                if isinstance(item, dict):
                    rows.append("; ".join(f"{k}: {v}" for k, v in item.items()
                                          if v not in (None, "", "null")))
                else:
                    rows.append(str(item))
            return "\n".join(r for r in rows if r)
        return str(value)
    return str(value)


def render_quotes(quotes) -> str:
    if not quotes:
        return ""
    return "\n".join(f"“{q}”" for q in quotes)


# --------------------------- load ---------------------------

def load_extraction(path: str) -> dict[str, dict[str, dict]]:
    """{record_id: {group_id: line}}."""
    out: dict[str, dict[str, dict]] = {}
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            r = json.loads(line)
            out.setdefault(str(r["record_id"]), {})[str(r["group_id"])] = r
    return out


def load_records(path: str) -> dict[str, dict]:
    out: dict[str, dict] = {}
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                r = json.loads(line)
                out[str(r["record_id"])] = r
    return out


def _quote_ok_against(quote: str, norm_haystack: str) -> bool:
    """Verbatim (fuzzy) check of one quote against a PRE-NORMALISED haystack."""
    for frag in parse_quote_fragments(quote):
        nfrag = norm(frag)
        if not nfrag:
            continue
        if nfrag in norm_haystack:
            continue
        if _fuzzy_in_haystack(nfrag, norm_haystack):
            continue
        return False
    return True


def validate_consensus_quotes(extraction: dict, records: dict) -> None:
    """Stamp `_quote_ok` on every reconciled field by re-checking its quotes
    against the source PDF text. The reconciler can merge or re-word quotes, so
    consensus quotes are validated here (in addition to per-extractor quotes) —
    this is what makes the workbook's quote-fail flags and the review queue's
    'unverified_quote' issue fire on the consensus a human actually reviews.
    A reported field with no quote is treated as unverified (ungrounded).

    The source haystack is normalised ONCE per study (not per quote) — the hot
    path across thousands of consensus quotes."""
    for rid, groups in extraction.items():
        rec = records.get(rid, {})
        haystack = norm(f"{rec.get('title', '')} {rec.get('abstract', '')} {rec.get('year', '')}")
        for line in groups.values():
            if line.get("gated_out"):
                continue
            fields = (line.get("reconciled") or {}).get("fields", {})
            for fld in fields.values():
                if not fld.get("reported"):
                    fld["_quote_ok"] = True
                    continue
                quotes = fld.get("quotes") or []
                fld["_quote_ok"] = bool(quotes) and all(
                    _quote_ok_against(q, haystack) for q in quotes)


# --------------------------- sheet builders ---------------------------

def _style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = WRAP
    ws.freeze_panes = "A2"


def _autosize(ws, widths: dict[int, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def build_index_sheet(wb, fw: Framework, extraction: dict, records: dict) -> None:
    ws = wb.active
    ws.title = "Index"
    headers = ["record_id", "title", "year", "groups_extracted", "gated_out_groups",
               "total_reported", "needs_human"]
    ws.append(headers)
    _style_header(ws, len(headers))
    for rid, groups in sorted(extraction.items()):
        rec = records.get(rid, {})
        extracted, gated, reported, human = [], [], 0, 0
        for gid, line in groups.items():
            if line.get("gated_out"):
                gated.append(gid)
                continue
            extracted.append(gid)
            audit = line.get("audit", {})
            reported += audit.get("n_reported", 0)
            human += audit.get("n_needs_human", 0)
        ws.append([rid, rec.get("title", ""), rec.get("year", ""),
                   ", ".join(extracted), ", ".join(gated), reported, human])
    for row in ws.iter_rows(min_row=2):
        row[1].alignment = WRAP
    _autosize(ws, {1: 14, 2: 60, 3: 6, 4: 22, 5: 16, 6: 12, 7: 12})


def build_group_sheet(wb, fw: Framework, gid: str, extraction: dict, records: dict) -> None:
    group = fw.group(gid)
    ws = wb.create_sheet(title=_safe_title(gid))
    headers = ["record_id", "title"]
    for f in group.fields:
        headers += [f"{f.id} value", f"{f.id} quotes", f"{f.id} loc", f"{f.id} agree"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for rid, groups in sorted(extraction.items()):
        line = groups.get(gid)
        if not line:
            continue
        rec = records.get(rid, {})
        row = [rid, rec.get("title", "")]
        if line.get("gated_out"):
            row.append("— gated out —")
            row += [""] * (len(headers) - 3)
            ws.append(row)
            continue
        fields = (line.get("reconciled") or {}).get("fields", {})
        for f in group.fields:
            fld = fields.get(f.id, {})
            row += [
                render_value(f, fld.get("value")) if fld.get("reported") else "",
                render_quotes(fld.get("quotes")),
                fld.get("location", ""),
                fld.get("agreement", ""),
            ]
        ws.append(row)
        # Highlight conflict / quote-fail cells on this row.
        r_idx = ws.max_row
        for i, f in enumerate(group.fields):
            fld = fields.get(f.id, {})
            base_col = 3 + i * 4  # value column for field i
            if fld.get("needs_human"):
                ws.cell(row=r_idx, column=base_col).fill = FILL_CONFLICT
            if fld.get("_quote_ok") is False:
                ws.cell(row=r_idx, column=base_col + 1).fill = FILL_QUOTEFAIL

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP
    widths = {1: 14, 2: 40}
    for i in range(len(group.fields)):
        widths[3 + i * 4] = 40
        widths[4 + i * 4] = 40
        widths[5 + i * 4] = 14
        widths[6 + i * 4] = 10
    _autosize(ws, widths)


def build_review_queue(wb, fw: Framework, extraction: dict, records: dict) -> int:
    ws = wb.create_sheet(title="Review queue")
    headers = ["record_id", "title", "group", "field_id", "field_label", "issue",
               "agreement", "consensus_value", "quotes", "notes"]
    ws.append(headers)
    _style_header(ws, len(headers))
    n = 0
    for rid, groups in sorted(extraction.items()):
        rec = records.get(rid, {})
        for gid, line in groups.items():
            if line.get("gated_out"):
                continue
            group = fw.group(gid)
            fields = (line.get("reconciled") or {}).get("fields", {})
            for f in group.fields:
                fld = fields.get(f.id, {})
                issues = []
                if fld.get("needs_human"):
                    issues.append("conflict/needs_human")
                if fld.get("_quote_ok") is False:
                    issues.append("unverified_quote")
                if not issues:
                    continue
                ws.append([rid, rec.get("title", ""), gid, f.id, f.label,
                           "; ".join(issues), fld.get("agreement", ""),
                           render_value(f, fld.get("value")),
                           render_quotes(fld.get("quotes")), fld.get("notes", "")])
                r_idx = ws.max_row
                ws.cell(row=r_idx, column=6).fill = FILL_CONFLICT
                n += 1
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP
    _autosize(ws, {1: 14, 2: 34, 3: 8, 4: 8, 5: 26, 6: 20, 7: 12, 8: 44, 9: 44, 10: 30})
    return n


def build_crosswalk_sheet(wb, fw: Framework, extraction: dict) -> None:
    ws = wb.create_sheet(title="Eval crosswalk")
    headers = ["eval_section", "field_id", "field_label", "#studies_reported"]
    ws.append(headers)
    _style_header(ws, len(headers))
    # Count studies with a reported value per field.
    reported_count: dict[str, int] = {}
    for rid, groups in extraction.items():
        for gid, line in groups.items():
            if line.get("gated_out"):
                continue
            fields = (line.get("reconciled") or {}).get("fields", {})
            for fid, fld in fields.items():
                if fld.get("reported"):
                    reported_count[fid] = reported_count.get(fid, 0) + 1
    for section, fids in fw.eval_crosswalk.items():
        for fid in fids:
            try:
                label = fw.field(fid).label
            except KeyError:
                label = "?"
            ws.append([section, fid, label, reported_count.get(fid, 0)])
    for row in ws.iter_rows(min_row=2):
        row[2].alignment = WRAP
    _autosize(ws, {1: 34, 2: 8, 3: 34, 4: 16})


def build_codebook_sheet(wb, fw: Framework) -> None:
    ws = wb.create_sheet(title="Codebook")
    headers = ["group", "field_id", "label", "type", "allowed", "what", "feeds_eval"]
    ws.append(headers)
    _style_header(ws, len(headers))
    for g in fw.groups:
        for f in g.fields:
            ws.append([g.id, f.id, f.label, f.type,
                       ", ".join(f.allowed), f.what, f.feeds])
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP
    _autosize(ws, {1: 10, 2: 8, 3: 30, 4: 16, 5: 30, 6: 70, 7: 34})


def _safe_title(gid: str) -> str:
    return str(gid)[:31]


# --------------------------- long JSONL ---------------------------

def write_long_jsonl(path: str, fw: Framework, extraction: dict, records: dict) -> int:
    # field -> eval sections (inverse crosswalk)
    field_to_sections: dict[str, list[str]] = {}
    for section, fids in fw.eval_crosswalk.items():
        for fid in fids:
            field_to_sections.setdefault(fid, []).append(section)

    n = 0
    with open(path, "w", encoding="utf-8") as out:
        for rid, groups in sorted(extraction.items()):
            rec = records.get(rid, {})
            for gid, line in groups.items():
                if line.get("gated_out"):
                    continue
                group = fw.group(gid)
                fields = (line.get("reconciled") or {}).get("fields", {})
                for f in group.fields:
                    fld = fields.get(f.id, {})
                    out.write(json.dumps({
                        "record_id": rid,
                        "title": rec.get("title", ""),
                        "year": rec.get("year", ""),
                        "group_id": gid,
                        "field_id": f.id,
                        "field_label": f.label,
                        "field_type": f.type,
                        "reported": bool(fld.get("reported")),
                        "value": fld.get("value"),
                        "value_str": render_value(f, fld.get("value")) if fld.get("reported") else "",
                        "quotes": fld.get("quotes", []),
                        "location": fld.get("location", ""),
                        "confidence": fld.get("confidence", ""),
                        "agreement": fld.get("agreement", ""),
                        "needs_human": bool(fld.get("needs_human")),
                        "quote_ok": fld.get("_quote_ok", None),
                        "eval_sections": field_to_sections.get(f.id, []),
                        "feeds": f.feeds,
                    }, ensure_ascii=False) + "\n")
                    n += 1
    return n


# --------------------------- main ---------------------------

def main():
    p = argparse.ArgumentParser(description="Export reconciled extractions to Excel + long JSONL.")
    p.add_argument("--project", default=config.DEFAULT_PROJECT,
                   choices=list(config.PROJECT_DIRS.keys()))
    p.add_argument("--framework", help="Override framework YAML path.")
    p.add_argument("--extraction", required=True, help="Extraction JSONL from extract.py.")
    p.add_argument("--records", required=True, help="records_<n>.jsonl (for titles).")
    p.add_argument("--out-xlsx", help="Output .xlsx (default: <extraction_dir>/reports/extraction.xlsx).")
    p.add_argument("--out-jsonl", help="Output long JSONL (default: alongside xlsx).")
    args = p.parse_args()

    fw = load_framework(args.framework or config.framework_path(args.project))
    extraction = load_extraction(args.extraction)
    records = load_records(args.records)
    validate_consensus_quotes(extraction, records)
    print(f"Loaded extraction for {len(extraction)} studies; framework {fw.ref}")

    reports_dir = config.extraction_dir(args.project) / "reports"
    reports_dir.mkdir(parents=True, exist_ok=True)
    out_xlsx = Path(args.out_xlsx) if args.out_xlsx else reports_dir / "extraction.xlsx"
    out_jsonl = Path(args.out_jsonl) if args.out_jsonl else reports_dir / "extraction_long.jsonl"

    # Long JSONL.
    n_rows = write_long_jsonl(str(out_jsonl), fw, extraction, records)
    print(f"Long-form: {n_rows} study×field rows → {out_jsonl}")

    # Workbook.
    wb = openpyxl.Workbook()
    build_index_sheet(wb, fw, extraction, records)
    for g in fw.groups:
        build_group_sheet(wb, fw, g.id, extraction, records)
    n_queue = build_review_queue(wb, fw, extraction, records)
    build_crosswalk_sheet(wb, fw, extraction)
    build_codebook_sheet(wb, fw)
    wb.save(out_xlsx)
    print(f"Workbook: {out_xlsx}")
    print(f"  Review queue: {n_queue} cells need a human (conflicts + unverified quotes)")


if __name__ == "__main__":
    main()
